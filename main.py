import threading
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
import tkinter.font as tkFont
import pandas as pd
import warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from pathlib import Path


class HackerrankLeaderboard:
    def __init__(self):
        self.prog_text = ''
        self.setup_root()
        self.create_widgets()

    def setup_root(self):
        self.root = tk.Tk()
        self.root.title("Hackerrank Leaderboard")
        self.root.configure(background='#404445')
        width = 1142
        height = 697
        screenwidth = self.root.winfo_screenwidth()
        screenheight = self.root.winfo_screenheight()
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.root.geometry(alignstr)
        self.root.resizable(width=False, height=False)

        # Ensure Leaderboards directory exists
        Path("Leaderboards").mkdir(exist_ok=True)

        # Set up window icon and protocol
        try:
            self.root.iconbitmap('venv/logo.ico')
        except:
            pass  # Skip if icon not found
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self):
        # Create the Enter ID label
        self.create_header_label()

        # Create the input field
        self.create_input_field()

        # Create the Generate and Combine buttons
        self.create_buttons()

    def create_header_label(self):
        ft = tkFont.Font(family='Helvetica', size=60, weight='bold')
        id_label = tk.Label(
            self.root,
            anchor="center",
            font=ft,
            fg="#FF6C40",
            justify="center",
            text="ENTER HACKERRANK ID'S!",
            bg='#404445'
        )
        id_label.place(x=15, y=2, width=1100, height=131)

    def create_input_field(self):
        self.entry = tk.Text(self.root)
        self.entry["borderwidth"] = "5px"
        self.entry['background'] = "black"
        ft = tkFont.Font(family='Times', size=25, weight="bold")
        self.entry["font"] = ft
        self.entry.insert('1.0', '   Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s')
        self.entry.bind("<FocusIn>", self.on_entry_click)
        self.entry["fg"] = "#FFE33E"
        self.entry["relief"] = "groove"
        self.entry.place(x=20, y=120, width=1101, height=431)
        self.entry["insertbackground"] = "#FFE33E"

    def create_buttons(self):
        # Generate button
        self.generate_btn = self.create_styled_button(
            "Generate Excel Sheets!",
            self.generate_sheets_command,
            "maroon",
            25,
            (60, 570, 500, 99)
        )

        # Combine button
        self.combine_btn = self.create_styled_button(
            "Combine Existing Excel Sheets",
            self.combine_excel_sheets,
            "#006400",
            25,
            (580, 570, 490, 99)
        )

    def create_styled_button(self, text, command, bg_color, font_size, placement):
        btn = tk.Button(self.root)
        btn.bind('<Enter>', lambda e: btn.config(background='black'))
        btn.bind('<Leave>', lambda e: btn.config(background=bg_color))
        btn.configure(
            background=bg_color,
            font=tkFont.Font(family='Times', size=font_size, weight='bold'),
            borderwidth="7px",
            fg="#FFE33E",
            justify="center",
            relief="groove",
            text=text,
            command=command
        )
        btn.place(x=placement[0], y=placement[1], width=placement[2], height=placement[3])
        return btn

    def create_progress_window(self, title="Please Wait..."):
        progress_window = tk.Toplevel(self.root)
        try:
            progress_window.iconbitmap('venv/logo.ico')
        except:
            pass
        progress_window.title(title)
        progress_window["borderwidth"] = "5px"
        progress_window["relief"] = "groove"
        progress_window.geometry("800x400")
        progress_window.resizable(False, False)
        progress_window['background'] = '#404445'

        # Configure progress text
        progress_text = tk.Text(progress_window, height=30, width=80)
        progress_text.configure(
            background="grey",
            fg='white',
            font=tkFont.Font(family='Times', size=20, weight='bold')
        )
        progress_text.pack(pady=80)

        # Configure progress bar
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TProgressbar",
                        thickness=20,
                        troughcolor='lightgrey',
                        background='#FF6C40')
        progress = ttk.Progressbar(progress_window, mode='determinate', style="TProgressbar")
        progress.place(x=50, y=10, width=700, height=50)

        return progress_window, progress_text, progress

    def generateExcelSheet(self, name, df):
        # This method now only handles the total leaderboard
        if name != 'TotalHackerrankLeaderBoard':
            return

        # Sort the DataFrame
        df = df.sort_values(by='Total Score', ascending=False)

        # Add rank after sorting
        df.insert(0, 'Rank', range(1, len(df) + 1))

        # Create Excel file
        filepath = Path(f'Leaderboards/{name}.xlsx')
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            self.apply_excel_formatting(writer.sheets['Sheet1'], df)

    def apply_excel_formatting(self, worksheet, df):
        # Define styles
        styles = {
            'header': {
                'font': Font(name='Arial', size=18, bold=True),
                'fill': PatternFill(start_color='00ADEAEA', end_color='00ADEAEA', fill_type='solid'),
            },
            'body': {
                'font': Font(name='Arial', size=14, bold=True),
                'fill': PatternFill(start_color='00C7ECEC', end_color='00C7ECEC', fill_type='solid'),
            },
            'common': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(bottom=Side(style='medium'))
            }
        }

        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Rank column
        for col in worksheet.columns:
            column = col[0].column_letter
            if column != 'A':
                worksheet.column_dimensions[column].width = 35

        # Set row height
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 25

        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.value = value
            self.apply_cell_style(cell, styles['header'], styles['common'])

        for row_num, row in enumerate(df.values, start=2):
            for col_num, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                self.apply_cell_style(cell, styles['body'], styles['common'])

    @staticmethod
    def apply_cell_style(cell, specific_style, common_style):
        for style_dict in (specific_style, common_style):
            for attr, value in style_dict.items():
                setattr(cell, attr, value)

    def fetch_hackerrank_data(self, tracker_name):
        data = []
        headers = {
            "User-agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 "
                          "Safari/537.36"
        }

        for offset in range(0, 1000, 100):
            url = f'https://www.hackerrank.com/rest/contests/{tracker_name}/leaderboard?offset={offset}&limit=100'
            try:
                response = requests.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                json_data = response.json()

                if not json_data.get('models'):
                    break

                for item in json_data['models']:
                    data.append({
                        'Name': item['hacker'],
                        'Score': item['score']
                    })

            except requests.RequestException as e:
                messagebox.showerror("Error", f"Failed to fetch data for {tracker_name}: {str(e)}")
                return None

        return pd.DataFrame(data) if data else None

    def generate_sheets_thread(self, tracker_names, progress_window, progress_text, progress):
        try:
            warnings.filterwarnings('ignore')
            all_participants = {}
            total_sheets = len(tracker_names)

            # Create a single workbook for all contest sheets
            contests_filepath = Path('Leaderboards/ContestLeaderboards.xlsx')
            with pd.ExcelWriter(contests_filepath, engine='openpyxl') as writer:
                for idx, tracker_name in enumerate(tracker_names, 1):
                    df = self.fetch_hackerrank_data(tracker_name)
                    if df is None:
                        continue

                    if df.empty:
                        messagebox.showinfo("Warning", f"{tracker_name} returned no data")
                        continue

                    # Update all_participants dictionary
                    for _, row in df.iterrows():
                        if row['Name'] not in all_participants:
                            all_participants[row['Name']] = {contest: 0 for contest in tracker_names}
                        all_participants[row['Name']][tracker_name] = row['Score']

                    # Sort the DataFrame
                    df = df.sort_values(by='Score', ascending=False)
                    # Add rank after sorting
                    df.insert(0, 'Rank', range(1, len(df) + 1))

                    # Write to the Excel file
                    df.to_excel(writer, sheet_name=tracker_name[:31],
                                index=False)  # Excel sheet names limited to 31 chars
                    self.apply_excel_formatting(writer.sheets[tracker_name[:31]], df)

                    # Update progress
                    self.update_progress(progress_window, progress_text, progress,
                                         f'\nFinished {tracker_name}!\n',
                                         int(idx / total_sheets * 100))

                # Generate total leaderboard in a separate file
                if all_participants:
                    self.generate_total_leaderboard(all_participants, tracker_names)
                    messagebox.showinfo("Success", "Sheets generated successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            self.cleanup_progress(progress_window)

    def generate_total_leaderboard(self, all_participants, tracker_names):
        total_data = []
        for participant, scores in all_participants.items():
            row = {'Name': participant}
            row.update(scores)
            row['Total Score'] = sum(scores.values())
            total_data.append(row)

        df_total = pd.DataFrame(total_data)
        columns = ['Name'] + tracker_names + ['Total Score']
        df_total = df_total[columns]
        self.generateExcelSheet('TotalHackerrankLeaderBoard', df_total)

    def update_progress(self, window, text_widget, progress_bar, message, value):
        text_widget.config(state=tk.NORMAL)
        text_widget.insert(tk.END, message)
        text_widget.see(tk.END)
        text_widget.config(state=tk.DISABLED)
        progress_bar['value'] = value
        window.update()

    def cleanup_progress(self, progress_window):
        self.root.attributes('-disabled', False)
        if progress_window.winfo_exists():
            progress_window.destroy()

    def generate_sheets_command(self):
        inp = self.entry.get(1.0, 'end-1c').strip()
        default_text = '   Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s'

        if inp == default_text or not inp:
            messagebox.showerror('Error', 'Please enter contest IDs!')
            return

        try:
            contest_ids = [id.strip() for id in inp.split(',') if id.strip()]
            if not contest_ids:
                messagebox.showerror('Error', 'No valid contest IDs entered!')
                return

            self.root.attributes('-disabled', True)
            progress_window, progress_text, progress = self.create_progress_window()

            threading.Thread(
                target=self.generate_sheets_thread,
                args=(contest_ids, progress_window, progress_text, progress),
                daemon=True
            ).start()

        except Exception as e:
            messagebox.showerror('Error', f'An error occurred: {str(e)}')
            self.root.attributes('-disabled', False)

    def combine_sheets_thread(self, student_file, hackerrank_file, progress_window, progress_text, progress):
        try:
            # Read student data file
            self.update_progress(progress_window, progress_text, progress, "Reading student data file...\n", 25)
            student_df = pd.read_excel(student_file)
            student_df = student_df[['Roll number', 'Hackerrank']].copy()
            student_df['Hackerrank'] = student_df['Hackerrank'].str.strip().str.lstrip('@').str.lower()

            # Read Hackerrank leaderboard file
            self.update_progress(progress_window, progress_text, progress, "Reading Hackerrank leaderboard file...\n",
                                 50)
            hackerrank_df = pd.read_excel(hackerrank_file)

            # Drop existing Rank and Total Score columns if they exist
            columns_to_drop = ['Rank', 'Total Score']
            hackerrank_df = hackerrank_df.drop(columns=[col for col in columns_to_drop if col in hackerrank_df.columns])

            # Clean data and convert to lowercase for matching
            self.update_progress(progress_window, progress_text, progress, "Processing data...\n", 75)
            student_df['Hackerrank'] = student_df['Hackerrank'].str.strip()
            hackerrank_df['Name'] = hackerrank_df['Name'].str.strip()
            hackerrank_df['Name_lower'] = hackerrank_df['Name'].str.lower()

            # Get score columns
            score_columns = [col for col in hackerrank_df.columns
                             if col not in ['Name', 'Name_lower', 'Rank', 'Total Score']]

            # Create matched sheet
            # First, create a DataFrame with all students and zero scores
            matched_df = student_df.copy()
            matched_df['Name'] = ''  # Empty column for Hackerrank names
            for col in score_columns:
                matched_df[col] = 0

            # Update scores for matched students
            for idx, student in matched_df.iterrows():
                match = hackerrank_df[hackerrank_df['Name_lower'] == student['Hackerrank']]
                if not match.empty:
                    matched_df.at[idx, 'Name'] = match.iloc[0]['Name']  # Use actual Hackerrank name
                    for col in score_columns:
                        matched_df.at[idx, col] = match.iloc[0][col]

            # Calculate total score for matched entries
            matched_df['Total Score'] = matched_df[score_columns].sum(axis=1)
            matched_df = matched_df.sort_values('Total Score', ascending=False)
            matched_df.insert(0, 'Rank', range(1, len(matched_df) + 1))

            # Create unmatched sheet
            # Get all Hackerrank usernames that weren't matched
            matched_usernames = matched_df[matched_df['Name'] != '']['Name'].str.lower()
            unmatched_hackerrank = hackerrank_df[~hackerrank_df['Name_lower'].isin(matched_usernames)].copy()

            # Prepare unmatched DataFrame
            unmatched_df = unmatched_hackerrank.drop('Name_lower', axis=1)
            unmatched_df['Roll number'] = ''  # Empty roll number for unmatched Hackerrank users
            unmatched_df['Total Score'] = unmatched_df[score_columns].sum(axis=1)
            unmatched_df = unmatched_df.sort_values('Total Score', ascending=False)
            unmatched_df.insert(0, 'Rank', range(1, len(unmatched_df) + 1))

            # Reorder columns for both dataframes
            final_cols = ['Rank', 'Roll number', 'Name'] + score_columns + ['Total Score']
            matched_df = matched_df[final_cols]
            unmatched_df = unmatched_df[final_cols]

            # Generate Excel file
            self.update_progress(progress_window, progress_text, progress, "Generating Excel file...\n", 90)
            with pd.ExcelWriter('Leaderboards/CombinedLeaderboard.xlsx', engine='openpyxl') as writer:
                matched_df.to_excel(writer, index=False, sheet_name='Matched Entries')
                self.apply_excel_formatting(writer.sheets['Matched Entries'], matched_df)

                unmatched_df.to_excel(writer, index=False, sheet_name='Unmatched Entries')
                self.apply_excel_formatting(writer.sheets['Unmatched Entries'], unmatched_df)

            # Prepare summary message
            students_with_scores = len(matched_df[matched_df['Name'] != ''])
            students_without_scores = len(matched_df[matched_df['Name'] == ''])
            unmatched_hackerrank_users = len(unmatched_df)

            messagebox.showinfo("Success",
                                f"Excel sheets generated successfully!\n\n"
                                f"Matched Entries Sheet:\n"
                                f"- Students with scores: {students_with_scores}\n"
                                f"- Students without participation: {students_without_scores}\n\n"
                                f"Unmatched Entries Sheet:\n"
                                f"- Unmatched Hackerrank users: {unmatched_hackerrank_users}\n\n"
                                f"Check both sheets in CombinedLeaderboard.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            self.cleanup_progress(progress_window)


    def combine_excel_sheets(self):
        # Show instruction message box
        messagebox.showinfo(
            "Instructions",
            "Please follow these steps:\n\n"
            "1. First, upload the Student Batch Excel sheet\n"
            "   (containing Roll Numbers and Hackerrank IDs)\n\n"
            "2. Then, upload the TotalHackerrankLeaderBoard.xlsx file\n"
            "   (generated from the previous step)"
        )

        try:
            student_file = filedialog.askopenfilename(
                title='Select Student Data Excel File',
                filetypes=[('Excel Files', '*.xlsx')],
                initialdir='Leaderboards/'
            )
            if not student_file:
                return

            hackerrank_file = filedialog.askopenfilename(
                title='Select Hackerrank Leaderboard Excel File',
                filetypes=[('Excel Files', '*.xlsx')],
                initialdir='Leaderboards/'
            )
            if not hackerrank_file:
                return

            self.root.attributes('-disabled', True)
            progress_window, progress_text, progress = self.create_progress_window("Combining Excel Sheets...")

            threading.Thread(
                target=self.combine_sheets_thread,
                args=(student_file, hackerrank_file, progress_window, progress_text, progress),
                daemon=True
            ).start()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.root.attributes('-disabled', False)

    def merge_dataframes(self, dataframes):
        # Merge all dataframes on Name column
        combined_df = dataframes[0]
        for df in dataframes[1:]:
            combined_df = pd.merge(combined_df, df, on='Name', how='outer')

        # Fill NaN values with 0
        combined_df = combined_df.fillna(0)

        # Add Total Score column
        score_columns = [col for col in combined_df.columns if col != 'Name']
        combined_df['Total Score'] = combined_df[score_columns].sum(axis=1)

        # Sort by Total Score
        return combined_df.sort_values('Total Score', ascending=False)

    def on_entry_click(self, event):
        if self.entry.get("1.0", 'end-1c').strip() == 'Enter Comma Separated values of HACKERRANK_CONTEST_ID\'s':
            self.entry.delete('1.0', tk.END)

    def on_closing(self):
        self.root.destroy()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = HackerrankLeaderboard()
    app.run()